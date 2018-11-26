import os
import openpyxl


class Exel:

    wb1 = openpyxl.load_workbook('Import_Folders_new.xlsx')
    sheet_names1 = wb1.sheetnames
    sheet1 = wb1[sheet_names1[1]]
    all_rows = sheet1.max_row
    all_column = sheet1.max_column

    wb2 = openpyxl.load_workbook('test.xlsx')
    sheet_names2 = wb2.sheetnames
    sheet2 = wb2[sheet_names2[0]]

    def create_file(self, file_name):
        # filename = "test.xlsx"
        directory_path = os.getcwd()
        file_path = directory_path + "\\" + file_name
        wn = openpyxl.Workbook()
        wn.save(file_path)

    def all_data(self)->list:
        names = []

        for i in range(1, self.all_rows+1):
            names.append([])

        for r in range(1, self.all_rows+1):
            for c in range(1, self.all_column+1):
                e = self.sheet1.cell(row=r, column=c)
                names[r-1].append(e.value)

        return names

    def one_column_values(self, column_name)->list:  # take string parameter of column letter

        column_list = []
        # ws = wb1.active
        # first_column = ws['B']
        #
        # for x in range(len(first_column)):
        #     column_list.append(first_column[x].value)
        #
        # return column_list

        value = 1

        for i in range(self.all_rows):
            column_list.append(self.sheet1[column_name + str(value)].value)
            value += 1

        return column_list

    def dictionary_values(self, column_one, column_two)->dict:  # take two string parameter of columns letters

        return dict(zip(self.one_column_values(column_one),
                        self.one_column_values(column_two)))

    def copy_all_data(self):
        for r in range(1, self.all_rows+1):
            for c in range(1, self.all_column+1):
                j = self.sheet2.cell(row=r, column=c)
                j.value = self.all_data()[r-1][c-1]

        self.wb2.save('test.xlsx')


temp = Exel()
